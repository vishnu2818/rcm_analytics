from django.db.models import Count
from django.contrib.auth import login
from .forms import *
from .models import *
from django.contrib.auth.models import User
from django.http import JsonResponse
from .forms import ExcelUploadForm
from .utils import *
from django.forms.models import model_to_dict
import openpyxl
from openpyxl.utils import get_column_letter
import csv
from django.http import HttpResponse
import difflib
import pandas as pd
from decimal import Decimal
from django.contrib import messages
from django.utils.dateparse import parse_date
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.template.loader import get_template
from xhtml2pdf import pisa
from collections import Counter, defaultdict
from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.http import HttpResponse
from django.forms.models import model_to_dict

import pandas as pd
import difflib
from .models import ExcelUpload, ExcelData
import pprint
import logging
from datetime import datetime, date
from django.shortcuts import render, redirect
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required
import pandas as pd
from .models import ExcelUpload, ExcelData  # adjust as needed

logger = logging.getLogger(__name__)

@login_required
def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        file_path = default_storage.save(f'temp/{file.name}', file)
        abs_path = default_storage.path(file_path)

        df = pd.read_excel(abs_path)
        row_count = len(df)
        columns = list(df.columns)

        # Save metadata
        excel_upload = ExcelUpload.objects.create(
            user=request.user,
            file_name=file.name,
            row_count=row_count,
            columns=columns
        )

        # Get all model field names
        model_fields = set(
            f.name for f in ExcelData._meta.get_fields() if f.concrete and not f.auto_created
        )

        for i, record in enumerate(df.to_dict(orient='records')):
            clean_record = {}

            for key, value in record.items():
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()
                elif pd.isna(value):
                    value = None

                if key in model_fields:
                    clean_record[key] = value
                    logger.debug(f"Row {i + 1}: {key} = {value} (type={type(value)})")

            try:
                ExcelData.objects.create(upload=excel_upload, **clean_record)
            except Exception as e:
                logger.error(f"❌ Error saving row {i + 1}: {e}")
                logger.error(f"Row data: {clean_record}")
                raise

        request.session['current_upload_id'] = excel_upload.id
        request.session['uploaded_exceldata_path'] = abs_path
        request.session['excel_headers'] = columns
        request.session['upload_flow'] = 'task'

        return redirect('map_task_fields1')

    return render(request, 'upload_excel_file.html')


# 2️⃣ Map Excel Columns to Model Fields
@login_required
def map_task_fields1(request):
    excel_headers = request.session.get('excel_headers')
    if not excel_headers:
        messages.error(request, "No headers found. Please upload a file.")
        return redirect('upload_task')

    model_fields = [f.name for f in ExcelData._meta.fields if f.name not in ['id', 'assigned_to']]

    header_mappings = []
    for header in excel_headers:
        header_key = header.lower().replace(" ", "_")
        match = difflib.get_close_matches(header_key, model_fields, n=1, cutoff=0.6)
        suggested = match[0] if match else ""
        header_mappings.append({
            'header': header,
            'suggested': suggested
        })

    if request.method == 'POST':
        field_mapping = {}
        for item in header_mappings:
            mapped_field = request.POST.get(item['header'])
            if mapped_field:
                field_mapping[item['header']] = mapped_field

        request.session['exceldata_field_mapping'] = field_mapping

        return redirect('excel_display_data_verbose')

    return render(request, 'map_excel_fields.html', {
        'header_mappings': header_mappings,
        'model_fields': model_fields,
    })


# 3️⃣ Display Data with Classification
@login_required
def excel_display_data_verbose(request):
    file_path = request.session.get('uploaded_exceldata_path')
    field_mapping = request.session.get('exceldata_field_mapping')
    upload_id = request.session.get('current_upload_id')

    if not file_path or not field_mapping or not upload_id:
        messages.error(request, "Missing session data. Please upload and map again.")
        return redirect('upload_task')

    # Step 1: Load Excel & Apply Field Mapping
    df = pd.read_excel(file_path)
    df = df.rename(columns=field_mapping)

    allowed_fields = [f.name for f in ExcelData._meta.fields if f.name not in ['id', 'assigned_to']]
    records = df[[col for col in df.columns if col in allowed_fields]].to_dict(orient='records')

    upload = get_object_or_404(ExcelUpload, id=upload_id)

    # Optional: Clear previous data for re-import
    ExcelData.objects.filter(upload=upload).delete()

    for record in records:
        ExcelData.objects.create(upload=upload, **record)

    # Step 2: Classify Data
    queryset = ExcelData.objects.filter(upload=upload)
    processed_data = []

    print("\n======= STARTING CLASSIFICATION PROCESS =======")

    for i, row in enumerate(queryset, 1):
        balance = row.balance_due or 0
        charge = row.net_charges or 0
        payments = row.payments or 0
        status = (row.status or '').lower()
        payor = (row.cur_pay_category or '').lower()
        pri_payor = (row.pri_payor_category or '').lower()
        schedule_track = (row.schedule_track or '').lower()

        print(f"\n--- Row {i} ---")

        # Payment Status
        if balance < 0:
            ps = "Negative balance"
        elif balance == 0 and charge == 0 and status in ['canceled', 'closed']:
            ps = "Canceled Trip"
        elif balance == 0 and charge > 0 and payments > 0:
            ps = "Paid & Closed"
        elif balance == 0 and payments == 0:
            ps = "Adjusted"
        elif payments > 0:
            ps = "Partially paid"
        else:
            ps = "Unpaid"

        # AR Status
        if ps == "Negative balance":
            ars = "Negative Ins AR"
        elif ps == "Canceled Trip":
            ars = "Canceled Trip"
        elif ps == "Paid & Closed" and pri_payor == 'patient' and payor == 'patient':
            ars = "Closed - Pt Pri"
        elif ps == "Paid & Closed":
            ars = "Closed - Ins Pri"
        elif ps == "Adjusted":
            ars = "Adjusted & Closed"
        elif payor == "patient" and "denials" not in schedule_track and "waystar" not in schedule_track:
            ars = "Open - Pt AR"
        else:
            ars = "Open - Ins AR"

        # Claim Status
        cs, cs_debug = classify_claim_status(row, ps, ars)
        for debug_step in cs_debug:
            print(f"  - {debug_step}")
        print(f"  - Final Claim Status: '{cs}'")

        row_data = model_to_dict(row)

        # Convert any datetime/date field to string (ISO)
        for key, value in row_data.items():
            if isinstance(value, (datetime, date)):
                row_data[key] = value.isoformat()
            elif value is None:
                row_data[key] = ""  # Ensure no None leaks to template

        row_data.update({
            'Payment Status': ps,
            'AR Status': ars,
            'Claim Status': cs,
        })
        processed_data.append(row_data)

    print("\n======= EXCEL CLASSIFICATION COMPLETE =======")
    print("Total Processed Rows:", len(processed_data))
    pprint.pprint(processed_data)

    return render(request, 'processed_excel_data.html', {
        'data': processed_data,
        'upload': upload,
    })


@login_required
def upload_excel1(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        file_path = default_storage.save(f'temp/{file.name}', file)
        abs_path = default_storage.path(file_path)

        # Read Excel
        df = pd.read_excel(abs_path)

        row_count = len(df)
        columns = list(df.columns)

        # Create ExcelUpload entry
        excel_upload = ExcelUpload.objects.create(
            user=request.user,
            file_name=file.name,
            row_count=row_count,
            columns=columns
        )

        # 🔥 FIXED: Set headers and path in session
        request.session['current_upload_id'] = excel_upload.id
        request.session['uploaded_exceldata_path'] = abs_path
        request.session['excel_headers'] = columns  # 🛠️ THIS LINE FIXES YOUR ISSUE
        request.session['upload_flow'] = 'excel'  # ✅ set the flow type
        return redirect('map_task_fields')  # Next step

    return render(request, 'upload_task_file.html')


# @login_required
# def confirm_excel_import(request):
#     file_path = request.session.get('uploaded_exceldata_path')
#     field_mapping = request.session.get('exceldata_field_mapping')
#     upload_id = request.session.get('current_upload_id')
#
#     if not file_path or not field_mapping or not upload_id:
#         return redirect('upload_excel')
#
#     excel_upload = get_object_or_404(ExcelUpload, id=upload_id)
#     df = pd.read_excel(file_path)
#     df = df.rename(columns=field_mapping)
#
#     allowed_fields = [f.name for f in ExcelData._meta.fields if f.name != 'id' and f.name != 'upload']
#     clean_df = df[[col for col in df.columns if col in allowed_fields]]
#     records = clean_df.to_dict(orient='records')
#
#     if request.method == 'POST':
#         for row in records:
#             ExcelData.objects.create(**row, upload=excel_upload)
#
#         messages.success(request, f"{len(records)} records imported successfully.")
#         return redirect('test-verbose')
#
#     return render(request, 'confirm_excel_import.html', {
#         'total_records': len(records),
#         'sample_preview': records[:5],
#     })


@login_required
def test_display_data_verbose(request):
    upload_id = request.session.get('current_upload_id')
    if not upload_id:
        messages.error(request, "Missing upload ID.")
        return redirect('upload_excel')

    try:
        upload = ExcelUpload.objects.get(id=upload_id, user=request.user)
    except ExcelUpload.DoesNotExist:
        return HttpResponse("❌ Upload not found or access denied", status=403)

    queryset = ExcelData.objects.filter(upload=upload)

    processed_data = []

    print("\n======= STARTING CLASSIFICATION PROCESS =======")

    for i, row in enumerate(queryset, 1):
        balance = row.balance_due or 0
        charge = row.net_charges or 0
        payments = row.payments or 0
        status = (row.status or '').lower()
        payor = (row.cur_pay_category or '').lower()
        pri_payor = (row.pri_payor_category or '').lower()
        schedule_track = (row.schedule_track or '').lower()

        print(f"\n--- Row {i} ---")

        # 1. Payment Status
        if balance < 0:
            ps = "Negative balance"
        elif balance == 0 and charge == 0 and status in ['canceled', 'closed']:
            ps = "Canceled Trip"
        elif balance == 0 and charge > 0 and payments > 0:
            ps = "Paid & Closed"
        elif balance == 0 and payments == 0:
            ps = "Adjusted"
        elif payments > 0:
            ps = "Partially paid"
        else:
            ps = "Unpaid"

        # 2. AR Status
        if ps == "Negative balance":
            ars = "Negative Ins AR"
        elif ps == "Canceled Trip":
            ars = "Canceled Trip"
        elif ps == "Paid & Closed" and pri_payor == 'patient' and payor == 'patient':
            ars = "Closed - Pt Pri"
        elif ps == "Paid & Closed":
            ars = "Closed - Ins Pri"
        elif ps == "Adjusted":
            ars = "Adjusted & Closed"
        elif payor == "patient" and "denials" not in schedule_track and "waystar" not in schedule_track:
            ars = "Open - Pt AR"
        else:
            ars = "Open - Ins AR"

        # 3. Claim Status
        cs, cs_debug = classify_claim_status(row, ps, ars)
        for debug_step in cs_debug:
            print(f"  - {debug_step}")
        print(f"  - Final Claim Status: '{cs}'")

        row_data = model_to_dict(row)

        # # ✅ FIX: Convert datetime/date objects to strings
        # for key, value in row_data.items():
        #     if isinstance(value, (datetime, date)):
        #         row_data[key] = value.isoformat()

        row_data.update({
            'Payment Status': ps,
            'AR Status': ars,
            'Claim Status': cs,
        })
        processed_data.append(row_data)

    print("\n======= CLASSIFICATION COMPLETE =======")

    return render(request, 'testing.html', {
        'data': processed_data,
        'upload': upload,
    })


@login_required
def home(request):
    uploads = ExcelUpload.objects.filter(user=request.user).order_by('-uploaded_at')
    return render(request, 'home.html', {'uploads': uploads})


def register_view(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            # Create user
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = User.objects.create_user(username=username, email=email, password=password)

            # Create profile manually (instead of relying on signal)
            Profile.objects.create(
                user=user,

                company_name=form.cleaned_data['company_name'],
                company_email=form.cleaned_data['company_email'],
                phone=form.cleaned_data['phone'],
                avg_claim_rate_per_month=form.cleaned_data['avg_claim_rate_per_month'],
                heard_about_us=form.cleaned_data['heard_about_us']
            )

            login(request, user)
            return redirect('home')
    else:
        form = UserRegistrationForm()

    return render(request, 'register.html', {'form': form})


@login_required
def start_chat(request, user_id):
    target_user = get_object_or_404(User, id=user_id)
    rooms = ChatRoom.objects.filter(users=request.user).filter(users=target_user)
    if rooms.exists():
        room = rooms.first()
    else:
        room = ChatRoom.objects.create()
        room.users.add(request.user, target_user)
    return redirect('chat_room', room_id=room.id)


@login_required
def chat_room(request, room_id):
    room = get_object_or_404(ChatRoom, id=room_id)
    if request.user not in room.users.all():
        return redirect('user_list')
    message = Message.objects.filter(room=room).order_by('timestamp')
    return render(request, 'chat_room.html', {'room': room, 'messages': message})


@login_required
def send_message(request):
    if request.method == 'POST':
        content = request.POST.get('content')
        room_id = request.POST.get('room_id')
        room = ChatRoom.objects.get(id=room_id)
        message = Message.objects.create(room=room, sender=request.user, content=content)
        return JsonResponse({'status': 'Message Sent', 'messages': message})
    return JsonResponse({'status': 'Failed'})


@login_required
def user_list(request):
    users = User.objects.exclude(id=request.user.id)
    return render(request, 'user_list.html', {'users': users})


@login_required
def upload_excel1(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = request.FILES['file']
                df = pd.read_excel(file, engine='openpyxl')

                # Create ExcelUpload record
                upload = ExcelUpload.objects.create(
                    user=request.user,
                    file_name=file.name,
                    row_count=len(df),
                    columns={col: str(df[col].dtype) for col in df.columns}
                )
                request.session['current_upload_id'] = upload.id

                serializable_data = [
                    {col: convert_to_serializable(row[col]) for col in df.columns}
                    for _, row in df.iterrows()
                ]

                created_objects = []
                for row_data in serializable_data:
                    try:
                        obj = ExcelData(
                            upload=upload,
                            company=row_data.get("Company", "Unknown"),
                            dos=parse_date(row_data.get("DOS")),
                            dosym=row_data.get("DOSYM", "Unknown"),
                            run_number=row_data.get("Run #", "Unknown"),
                            inc_number=row_data.get("Inc #t", "Unknown"),
                            customer=row_data.get("Cust.", None),
                            dob=row_data.get("DOB", None),
                            status=row_data.get("Status", "Unknown"),
                            prim_pay=row_data.get("Prim Pay", "Unknown"),
                            pri_payor_category=row_data.get("Pri Payor Category", "Unknown"),
                            cur_pay=row_data.get("Cur Pay", "Unknown"),
                            cur_pay_category=row_data.get("Cur Pay Category", "Unknown"),
                            schedule_track=row_data.get("Schedule/Track", "Unknown"),
                            event_step=row_data.get("Event/Step", "Unknown"),
                            coll=row_data.get("Coll", "NO"),
                            gross_charges=parse_decimal(row_data.get("Gross Charges")),
                            contr_allow=parse_decimal(row_data.get("Contr Allow")),
                            net_charges=parse_decimal(row_data.get("Net Charges")),
                            revenue_adjustments=parse_decimal(row_data.get("Revenue Adjustments")),
                            payments=parse_decimal(row_data.get("Payments")),
                            write_offs=parse_decimal(row_data.get("Write-Offs")),
                            refunds=parse_decimal(row_data.get("Refunds")),
                            balance_due=parse_decimal(row_data.get("Balance Due")),
                            aging_date=parse_date(row_data.get("Aging Date")),
                            last_event_date=parse_date(row_data.get("Last Event Date")),
                            ordering_facility=row_data.get("Ordering Facility", None),
                            vehicle=str(row_data.get("Vehicle", "Unknown")),
                            call_type=row_data.get("Call Type", "Unknown"),
                            priority=row_data.get("Priority", "Unknown"),
                            call_type_priority=row_data.get("Call Type - Priority", "Unknown"),
                            primary_icd=row_data.get("Primary ICD", "Unknown"),
                            loaded_miles=parse_decimal(row_data.get("Loaded Miles", 0.0)),
                            pickup_facility=row_data.get("Pickup Facility", None),
                            pickup_modifier=row_data.get("Pickup Modifier", "Unknown"),
                            pickup_address=row_data.get("Pickup Address", None),
                            pickup_city=row_data.get("Pickup City", "Unknown"),
                            pickup_state=row_data.get("Pickup State", "NA"),
                            pickup_zip=str(row_data.get("Pickup Zip", "00000")),
                            dropoff_facility=row_data.get("DropOff Facility", "Unknown"),
                            dropoff_modifier=row_data.get("DropOff Modifier", "Unknown"),
                            dropoff_address=row_data.get("DropOff Address", None),
                            dropoff_city=row_data.get("DropOff City", "Unknown"),
                            dropoff_state=row_data.get("DropOff State", "NA"),
                            dropoff_zip=str(row_data.get("DropOff Zip", "00000")),
                            import_date=parse_date(row_data.get("Import Date")),
                            import_date_ym=row_data.get("Import Date YM", "Unknown"),
                            med_nec=row_data.get("Med Nec", "Unknown"),
                            accident_type=row_data.get("Accident Type", None),
                            assigned_group=str(row_data.get("Assigned Group", None)),
                            location=row_data.get("Location", "Unknown"),
                            last_modified_date=parse_date(row_data.get("Last Modified Date")),
                            last_modified_by=row_data.get("Last Modified By", "Unknown"),
                            team=row_data.get("Team", "Unknown"),
                            job=row_data.get("Job", "Unknown"),
                            emsmart_id=row_data.get("EMSmartID", "Unknown"),
                            prior_auth=row_data.get("Prior Auth", None),
                        )
                        created_objects.append(obj)
                    except Exception as row_error:
                        print(f"❌ Error: {row_error}")

                ExcelData.objects.bulk_create(created_objects)
                messages.success(request, f"✅ Uploaded {file.name} with {len(created_objects)} records.")
                return redirect('test-verbose')

            except Exception as e:
                messages.error(request, f"❌ Upload failed: {str(e)}")

    else:
        form = ExcelUploadForm()

    return render(request, 'upload.html', {'form': form})


def download_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExcelData Preview"

    rows = ExcelData.objects.filter(upload__user=request.user)[:50]

    data = []
    for row in rows:
        ps = classify_payment_status(row)
        ars = classify_ar_status(row, ps)

        row_dict = {
            'Company': row.company,
            'Date of Service': row.dos,
            'DOS YM': row.dosym,
            'Run Number': row.run_number,
            'Incident Number': row.inc_number,
            'Customer': row.customer,
            'DOB': row.dob,
            'Status': row.status,
            'Primary Payor': row.prim_pay,
            'Primary Payor Category': row.pri_payor_category,
            'Current Payor': row.cur_pay,
            'Current Payor Category': row.cur_pay_category,
            'Schedule/Track': row.schedule_track,
            'Event Step': row.event_step,
            'COLL': row.coll,
            'Gross Charges': float(row.gross_charges),
            'Contractual Allowance': float(row.contr_allow),
            'Net Charges': float(row.net_charges),
            'Revenue Adjustments': float(row.revenue_adjustments),
            'Payments': float(row.payments),
            'Write-offs': float(row.write_offs),
            'Refunds': float(row.refunds),
            'Balance Due': float(row.balance_due),
            'Aging Date': row.aging_date,
            'Last Event Date': row.last_event_date,
            'Ordering Facility': row.ordering_facility,
            'Vehicle': row.vehicle,
            'Call Type': row.call_type,
            'Priority': row.priority,
            'Call Type Priority': row.call_type_priority,
            'Primary ICD': row.primary_icd,
            'Loaded Miles': float(row.loaded_miles),
            'Pickup Facility': row.pickup_facility,
            'Pickup Modifier': row.pickup_modifier,
            'Pickup Address': row.pickup_address,
            'Pickup City': row.pickup_city,
            'Pickup State': row.pickup_state,
            'Pickup ZIP': row.pickup_zip,
            'Dropoff Facility': row.dropoff_facility,
            'Dropoff Modifier': row.dropoff_modifier,
            'Dropoff Address': row.dropoff_address,
            'Dropoff City': row.dropoff_city,
            'Dropoff State': row.dropoff_state,
            'Dropoff ZIP': row.dropoff_zip,
            'Import Date': row.import_date,
            'Import Date YM': row.import_date_ym,
            'Medical Necessity': row.med_nec,
            'Accident Type': row.accident_type,
            'Assigned Group': row.assigned_group,
            'Location': row.location,
            'Last Modified Date': row.last_modified_date,
            'Last Modified By': row.last_modified_by,
            'Team': row.team,
            'Job': row.job,
            'EMSmart ID': row.emsmart_id,
            'Prior Auth': row.prior_auth,
            'Payment Status': ps,
            'AR Status': ars,
        }

        data.append(row_dict)

    if not data:
        return HttpResponse("No data to export.", status=400)

    # Write headers
    headers = list(data[0].keys())
    ws.append(headers)

    # Write rows
    for row in data:
        ws.append([row.get(col, '') for col in headers])

    # Adjust column widths
    for col_num, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=excel_data_preview.xlsx'
    wb.save(response)
    return response


def download_pdf(request):
    rows = ExcelData.objects.filter(upload__user=request.user)[:50]

    data = []
    for row in rows:
        ps = classify_payment_status(row)
        ars = classify_ar_status(row, ps)

        row_dict = {
            'Company': row.company,
            'Date of Service': row.dos,
            'DOS YM': row.dosym,
            'Run Number': row.run_number,
            'Incident Number': row.inc_number,
            'Customer': row.customer,
            'DOB': row.dob,
            'Status': row.status,
            'Primary Payor': row.prim_pay,
            'Primary Payor Category': row.pri_payor_category,
            'Current Payor': row.cur_pay,
            'Current Payor Category': row.cur_pay_category,
            'Schedule/Track': row.schedule_track,
            'Event Step': row.event_step,
            'COLL': row.coll,
            'Gross Charges': float(row.gross_charges),
            'Contractual Allowance': float(row.contr_allow),
            'Net Charges': float(row.net_charges),
            'Revenue Adjustments': float(row.revenue_adjustments),
            'Payments': float(row.payments),
            'Write-offs': float(row.write_offs),
            'Refunds': float(row.refunds),
            'Balance Due': float(row.balance_due),
            'Aging Date': row.aging_date,
            'Last Event Date': row.last_event_date,
            'Ordering Facility': row.ordering_facility,
            'Vehicle': row.vehicle,
            'Call Type': row.call_type,
            'Priority': row.priority,
            'Call Type Priority': row.call_type_priority,
            'Primary ICD': row.primary_icd,
            'Loaded Miles': float(row.loaded_miles),
            'Pickup Facility': row.pickup_facility,
            'Pickup Modifier': row.pickup_modifier,
            'Pickup Address': row.pickup_address,
            'Pickup City': row.pickup_city,
            'Pickup State': row.pickup_state,
            'Pickup ZIP': row.pickup_zip,
            'Dropoff Facility': row.dropoff_facility,
            'Dropoff Modifier': row.dropoff_modifier,
            'Dropoff Address': row.dropoff_address,
            'Dropoff City': row.dropoff_city,
            'Dropoff State': row.dropoff_state,
            'Dropoff ZIP': row.dropoff_zip,
            'Import Date': row.import_date,
            'Import Date YM': row.import_date_ym,
            'Medical Necessity': row.med_nec,
            'Accident Type': row.accident_type,
            'Assigned Group': row.assigned_group,
            'Location': row.location,
            'Last Modified Date': row.last_modified_date,
            'Last Modified By': row.last_modified_by,
            'Team': row.team,
            'Job': row.job,
            'EMSmart ID': row.emsmart_id,
            'Prior Auth': row.prior_auth,
            'Payment Status': ps,
            'AR Status': ars,
        }

        data.append(row_dict)

    if not data:
        return HttpResponse("No data available for PDF generation", status=400)

    template = get_template('pdf_template.html')
    html = template.render({'data': data})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="excel_data_preview.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response


def dashboard_view(request):
    rows = ExcelData.objects.filter(upload__user=request.user)

    processed = []
    net_charges = 0
    total_payments = 0
    total_balance = 0
    ar_days = []

    for row in rows:
        # Accessing direct model fields
        ps = classify_payment_status(row)
        ars = classify_ar_status(row, ps)

        net = float(getattr(row, 'net_charges', 0))
        paid = float(getattr(row, 'payments', 0))
        balance = float(getattr(row, 'balance_due', 0))

        net_charges += net
        total_payments += paid
        total_balance += balance

        # Aging Days
        aging_str = getattr(row, 'aging_date', None)
        if aging_str:
            try:
                aging_date = datetime.strptime(str(aging_str), "%Y-%m-%d")
                ar_days.append((datetime.now() - aging_date).days)
            except Exception:
                pass

        # Build dict of all fields for display
        d = {
            'Patient Name': getattr(row, 'patient_name', ''),
            'Net Charges': net,
            'Payments': paid,
            'Balance Due': balance,
            'Status': getattr(row, 'status', ''),
            'Current Payor Category': getattr(row, 'cur_pay_category', ''),
            'Primary Payor Category': getattr(row, 'pri_payor_category', ''),
            'Schedule/Track': getattr(row, 'schedule_track', ''),
            'Import Date YM': getattr(row, 'import_date_ym', ''),
            'Aging Date': aging_str,
            'Payment Status': ps,
            'AR Status': ars,
        }

        processed.append(d)

    # Aggregations
    payment_counts = Counter(d['Payment Status'] for d in processed)
    ar_counts = Counter(d['AR Status'] for d in processed)
    payor_counts = Counter(d.get('Primary Payor Category', 'Unknown') for d in processed)
    track_counts = Counter(d.get('Schedule/Track', 'Unknown') for d in processed)

    import_trends = defaultdict(lambda: {'charges': 0, 'payments': 0})
    for d in processed:
        ym = d.get('Import Date YM')
        import_trends[ym]['charges'] += float(d.get('Net Charges', 0))
        import_trends[ym]['payments'] += float(d.get('Payments', 0))

    avg_ar_days = round(sum(ar_days) / len(ar_days), 1) if ar_days else 0

    return render(request, 'dashboard.html', {
        'total_claims': len(processed),
        'net_charges': net_charges,
        'total_payments': total_payments,
        'total_balance': total_balance,
        'avg_ar_days': avg_ar_days,
        'payment_counts': dict(payment_counts),
        'ar_counts': dict(ar_counts),
        'payor_counts': dict(payor_counts),
        'track_counts': dict(track_counts),
        'import_trends': dict(import_trends),
    })


from django.contrib.auth.decorators import login_required


@login_required
def upload_task_file(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        file_path = default_storage.save(f'temp/{file.name}', file)
        abs_path = default_storage.path(file_path)

        # Read Excel
        df = pd.read_excel(abs_path)
        row_count = len(df)
        columns = list(df.columns)

        # Create ExcelUpload entry
        excel_upload = ExcelUpload.objects.create(
            user=request.user,
            file_name=file.name,
            row_count=row_count,
            columns=columns
        )

        # 🔥 FIXED: Set headers and path in session
        request.session['current_upload_id'] = excel_upload.id
        request.session['uploaded_exceldata_path'] = abs_path
        request.session['excel_headers'] = columns  # 🛠️ THIS LINE FIXES YOUR ISSUE
        request.session['upload_flow'] = 'task'  # ✅ set the flow type
        return redirect('map_task_fields')  # Next step

    return render(request, 'upload_task_file.html')


@login_required
def map_task_fields(request):
    excel_headers = request.session.get('excel_headers')
    if not excel_headers:
        messages.error(request, "No headers found. Please upload a file.")
        return redirect('upload_task')

    model_fields = [f.name for f in ExcelData._meta.fields if f.name not in ['id', 'assigned_to']]

    header_mappings = []
    for header in excel_headers:
        header_key = header.lower().replace(" ", "_")
        match = difflib.get_close_matches(header_key, model_fields, n=1, cutoff=0.6)
        suggested = match[0] if match else ""
        header_mappings.append({
            'header': header,
            'suggested': suggested
        })

    if request.method == 'POST':
        field_mapping = {}
        for item in header_mappings:
            mapped_field = request.POST.get(item['header'])
            if mapped_field:
                field_mapping[item['header']] = mapped_field

        request.session['exceldata_field_mapping'] = field_mapping

        # ✅ Decide which confirm view to go to
        flow = request.session.get('upload_flow')
        if flow == 'task':
            return redirect('confirm_exceldata_import')
        else:
            return redirect('test-verbose')

    return render(request, 'map_task_fields.html', {
        'header_mappings': header_mappings,
        'model_fields': model_fields,
    })


@login_required
def confirm_exceldata_import(request):
    file_path = request.session.get('uploaded_exceldata_path')
    field_mapping = request.session.get('exceldata_field_mapping')
    upload_id = request.session.get('current_upload_id')  # get saved upload id

    if not file_path or not field_mapping or not upload_id:
        return redirect('upload_task')

    excel_upload = get_object_or_404(ExcelUpload, id=upload_id)

    df = pd.read_excel(file_path)
    df = df.rename(columns=field_mapping)

    allowed_fields = [f.name for f in ExcelData._meta.fields if f.name not in ['id', 'assigned_to']]
    data_to_import = df[[col for col in df.columns if col in allowed_fields]]
    task_pool = data_to_import.to_dict(orient='records')

    employees = Employee.objects.all()

    if request.method == 'POST':
        # Get counts assigned manually per employee from form
        custom_assignments = {}
        total_requested = 0
        for emp in employees:
            count = int(request.POST.get(f'emp_{emp.id}', 0))
            custom_assignments[emp] = count
            total_requested += count

        if total_requested > len(task_pool):
            messages.error(request, "Assigned task count exceeds total available tasks.")
            return redirect('confirm_exceldata_import')

        # Assign tasks accordingly
        assigned_tasks = []
        assigned_index = 0
        for emp, count in custom_assignments.items():
            for _ in range(count):
                if assigned_index >= len(task_pool):
                    break
                row_data = task_pool[assigned_index]
                assigned_tasks.append((row_data, emp))
                assigned_index += 1

        # Save to DB
        date_fields = ['dos', 'aging_date', 'last_event_date', 'import_date', 'last_modified_date']
        for row_data, emp in assigned_tasks:
            row_kwargs = {}
            for field in allowed_fields:
                value = row_data.get(field)
                if field in date_fields:
                    value = safe_date(value)
                row_kwargs[field] = value

            row_kwargs.pop('upload', None)
            ExcelData.objects.create(**row_kwargs, upload=excel_upload, assigned_to=emp)

        messages.success(request, f"{len(assigned_tasks)} rows imported & assigned successfully.")
        return redirect('employee_target_list')

    # Prepare preview data: calculate effective target count by ramp% for each employee
    preview_data = []
    for emp in employees:
        target = emp.target or Decimal(0)
        ramp = Decimal(emp.ramp_percent or 0)
        effective = int(target * ramp / 100)
        preview_data.append({
            'id': emp.id,
            'name': emp.employee_name,
            'target': target,
            'ramp': ramp,
            'effective': effective,
        })

    return render(request, 'preview_task_assignment.html', {
        'preview_data': preview_data,
        'total_tasks': len(task_pool),
    })


@login_required
def dashboard(request):
    # Summarize task count per employee
    summary = ExcelData.objects.values('assigned_to__employee_name').annotate(task_count=models.Count('id')).order_by(
        '-task_count')

    return render(request, 'dashboard.html', {
        'summary': summary
    })


# Employee CURD Operations

@login_required
def employee_target_list(request):
    employees = Employee.objects.select_related('client_name')

    # Attach task info
    for emp in employees:
        emp.task_count = ExcelData.objects.filter(assigned_to=emp).count()
        emp.tasks = ExcelData.objects.filter(assigned_to=emp).order_by('-id')[:5]

    # Export CSV
    if 'export' in request.GET:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="employee_targets.csv"'

        writer = csv.writer(response)
        writer.writerow(['Employee Name', 'Client Name', 'Target', 'Ramp %', 'Task Count', 'Recent Task IDs'])

        for emp in employees:
            recent_task_ids = ', '.join([task.emsmart_id or 'N/A' for task in emp.tasks])
            writer.writerow([
                emp.employee_name,
                emp.client_name.name if emp.client_name else '',
                emp.target,
                emp.ramp_percent,
                emp.task_count,
                recent_task_ids
            ])
        return response

    return render(request, 'employee_target_list.html', {'targets': employees})


@login_required
def employee_target_create(request):
    form = EmployeeForm(request.POST or None)
    if form.is_valid():
        employee = form.save(commit=False)
        employee.created_by = request.user
        employee.save()
        form.save_m2m()  # For regular m2m fields (not needed here, but good habit)

        # Save task assignments
        tasks = form.cleaned_data['tasks']
        tasks.update(assigned_to=employee)

        return redirect('employee_target_list')
    return render(request, 'employee_target_form.html', {'form': form, 'action': 'Create'})


from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render, redirect
from .models import Employee, ExcelData
from .forms import EmployeeForm


@login_required
def employee_target_update(request, pk):
    target = get_object_or_404(Employee, pk=pk)
    form = EmployeeForm(request.POST or None, instance=target)

    if form.is_valid():
        employee = form.save()

        # Remove previously assigned tasks from this employee
        ExcelData.objects.filter(assigned_to=employee).update(assigned_to=None)

        # Assign newly selected tasks
        tasks = form.cleaned_data.get('tasks')
        if tasks:
            tasks.update(assigned_to=employee)

        return redirect('employee_target_list')

    # Pre-populate the task selection
    form.fields['tasks'].initial = ExcelData.objects.filter(assigned_to=target)

    return render(request, 'employee_target_form.html', {'form': form, 'action': 'Update'})


@login_required
def employee_target_delete(request, pk):
    target = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        target.delete()
        return redirect('employee_target_list')
    return render(request, 'employee_target_confirm_delete.html', {'target': target})


@login_required
def employee_target_dashboard(request):
    # Example dashboard statistics
    stats = Employee.objects.aggregate(
        total_targets=Count('id'),
        average_ramp=models.Avg('ramp_percent'),
    )
    return render(request, 'employee_target_dashboard.html', {'stats': stats})


@login_required
def qa_audit_list(request):
    audits = QAAudit.objects.select_related('claim', 'audited_by').all()
    return render(request, 'qa_audit_list.html', {'audits': audits})


@login_required
def qa_audit_create(request):
    if request.method == 'POST':
        form = QAAuditForm(request.POST)
        if form.is_valid():
            audit = form.save(commit=False)

            # Get first client as default (you can change this logic)
            client = Client.objects.first()

            # Create or get the employee linked to the logged-in user
            employee, created = Employee.objects.get_or_create(
                email=request.user.email,
                defaults={
                    'employee_name': request.user.get_full_name() or request.user.username,
                    'client_name': client,
                    'target': 100,  # Default target
                    'ramp_percent': 0.0,
                    'created_by': request.user,
                }
            )

            audit.audited_by = employee
            audit.save()
            messages.success(request, "✅ Audit submitted successfully.")
            return redirect('qa_audit_list')
    else:
        form = QAAuditForm()

    return render(request, 'qa_audit_form.html', {'form': form})


def safe_date(value):
    if isinstance(value, datetime):
        return value.date()
    elif isinstance(value, str):
        return parse_date(value)
    return None


# def auto_map_headers(excel_headers, model_fields):
#     mapping = {}
#     for header in excel_headers:
#         key = header.strip().lower().replace(" ", "_")
#         match = difflib.get_close_matches(key, model_fields, n=1, cutoff=0.6)
#         if match:
#             mapping[header] = match[0]
#     return mapping


# @login_required
# def upload_excel_with_automap(request):
#     if request.method == 'POST' and request.FILES.get('excel_file'):
#         file = request.FILES['excel_file']
#         file_path = default_storage.save(f'temp/{file.name}', file)
#         abs_path = default_storage.path(file_path)
#
#         try:
#             df = pd.read_excel(abs_path)
#             columns = list(df.columns)
#
#             # Save ExcelUpload entry
#             upload = ExcelUpload.objects.create(
#                 user=request.user,
#                 file_name=file.name,
#                 row_count=len(df),
#                 columns=columns
#             )
#
#             request.session['current_upload_id'] = upload.id
#
#             # Get model fields (excluding some)
#             allowed_fields = [f.name for f in ExcelData._meta.fields if f.name not in ['id', 'assigned_to']]
#
#             # Auto-map Excel headers to model fields
#             header_mapping = auto_map_headers(columns, allowed_fields)
#
#             # Rename columns and filter only valid fields
#             df = df.rename(columns=header_mapping)
#             df = df[[col for col in df.columns if col in allowed_fields]]
#
#             # Prepare and save ExcelData records
#             date_fields = ['dos', 'aging_date', 'last_event_date', 'import_date', 'last_modified_date']
#             decimal_fields = ['gross_charges', 'net_charges', 'payments', 'balance_due']
#
#             rows = []
#             for _, row in df.iterrows():
#                 data = {}
#                 for field in df.columns:
#                     value = row.get(field)
#                     if field in date_fields:
#                         value = safe_date(value)
#                     elif field in decimal_fields:
#                         value = parse_decimal(value)
#                     data[field] = value
#                 rows.append(ExcelData(**data, upload=upload))
#
#             ExcelData.objects.bulk_create(rows)
#
#             messages.success(request, f"Uploaded and imported {len(rows)} rows successfully.")
#             return redirect(f'/test-verbose/?upload_id={upload.id}')
#
#         except Exception as e:
#             messages.error(request, f"Upload failed: {e}")
#
#     return render(request, 'upload_auto.html')
#

# excel update & delete code

@login_required
def edit_exceldata(request, pk):
    instance = get_object_or_404(ExcelData, id=pk)
    if request.method == 'POST':
        form = ExcelDataForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('test-verbose')
        else:
            print(form.errors)  # Debugging
    else:
        form = ExcelDataForm(instance=instance)
    return render(request, 'edit_exceldata.html', {'form': form})


@login_required
def delete_exceldata(request, pk):
    instance = get_object_or_404(ExcelData, id=pk)
    if request.method == 'POST':
        instance.delete()
        return redirect('test-verbose')
    return render(request, 'confirm_delete.html', {'object': instance})
